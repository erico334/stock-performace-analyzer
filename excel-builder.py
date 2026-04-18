"""
excel_builder.py
Builds the full multi-sheet Excel report.
"""
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import analyzer as az

FONT_NAME = "Arial"
DARK    = "1F3864"
MID_BLU = "2E75B6"
TOT_BG  = "D6DCE4"

thin = Side(style="thin", color="CCCCCC")
bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

def fl(h):
    return PatternFill("solid", fgColor=h.lstrip("#"))

def fn(bold=False, sz=10, col="1F1F1F"):
    return Font(name=FONT_NAME, bold=bold, size=sz, color=col)

def hf(col="FFFFFF", sz=10):
    return Font(name=FONT_NAME, bold=True, size=sz, color=col)

def ca(): return Alignment(horizontal="center", vertical="center")
def la(): return Alignment(horizontal="left",   vertical="center")

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def write_title(ws, text, subtitle, ncols=8):
    end = get_column_letter(ncols + 1)
    ws.merge_cells(f"B1:{end}1")
    ws["B1"] = text
    ws["B1"].font = hf(sz=13); ws["B1"].fill = fl(DARK); ws["B1"].alignment = la()
    ws.row_dimensions[1].height = 30
    ws.merge_cells(f"B2:{end}2")
    ws["B2"] = subtitle
    ws["B2"].font = Font(name=FONT_NAME, size=10, italic=True, color="FFFFFF")
    ws["B2"].fill = fl("2E4057"); ws["B2"].alignment = la()
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 8

def write_metric_cards(ws, metrics, start_row=4, bg="F5F5F5"):
    for i, (lbl, val, fmt) in enumerate(metrics):
        sc = i * 2 + 2; ec = sc + 1
        ws.merge_cells(start_row=start_row,   start_column=sc, end_row=start_row,   end_column=ec)
        ws.merge_cells(start_row=start_row+1, start_column=sc, end_row=start_row+1, end_column=ec)
        c1 = ws.cell(start_row, sc)
        c1.value = lbl; c1.font = fn(bold=True, sz=9, col="595959")
        c1.fill = fl(bg); c1.alignment = ca(); c1.border = bdr
        c2 = ws.cell(start_row+1, sc)
        c2.value = val; c2.font = Font(name=FONT_NAME, size=13, bold=True, color=DARK)
        c2.fill = fl("FFFFFF"); c2.alignment = ca(); c2.border = bdr
        if fmt: c2.number_format = fmt
    ws.row_dimensions[start_row].height   = 18
    ws.row_dimensions[start_row+1].height = 24
    ws.row_dimensions[start_row+2].height = 8

def write_table_headers(ws, headers, row, bg=DARK, fg="FFFFFF", height=20):
    for i, h in enumerate(headers, 2):
        c = ws.cell(row, i); c.value = h
        c.font = hf(col=fg); c.fill = fl(bg); c.alignment = ca(); c.border = bdr
    ws.row_dimensions[row].height = height

def write_totals_row(ws, row, ncols, sum_cols, bg=TOT_BG, data_start=8):
    for ci in range(2, ncols + 2):
        c = ws.cell(row, ci)
        if ci in sum_cols:
            c.value = f"=SUM({get_column_letter(ci)}{data_start}:{get_column_letter(ci)}{row-1})"
            c.number_format = "#,##0"
        elif ci == 2:
            c.value = "TOTAL"
        else:
            c.value = ""
        c.font = hf(col="1F1F1F"); c.fill = fl(bg); c.border = bdr; c.alignment = ca()
    ws.row_dimensions[row].height = 20


def _row_bg(days, idx):
    if pd.isna(days): return "F2F2F2" if idx % 2 == 0 else "FFFFFF"
    d = int(days)
    if d >= 366: return "D9534F"
    if d >= 271: return "EAAAA0" if idx % 2 == 0 else "F5C6C0"
    if d >= 181: return "F4CCCC" if idx % 2 == 0 else "FAE0E0"
    if d >= 121: return "F8CBAD" if idx % 2 == 0 else "FDDCCA"
    if d >= 91:  return "FCE4D6" if idx % 2 == 0 else "FDF0E8"
    if d >= 61:  return "FFEB9C" if idx % 2 == 0 else "FFFDE8"
    if d >= 31:  return "FFF2CC" if idx % 2 == 0 else "FFFDF0"
    if d >= 15:  return "F2F8EC" if idx % 2 == 0 else "FAFFF5"
    if d >= 8:   return "EBF5E0" if idx % 2 == 0 else "F5FCF0"
    return "E2EFDA" if idx % 2 == 0 else "EFFAF0"


def _flag(row):
    days  = row.get("DAYS_SINCE_SALE")
    qty_h = row.get("QTY AT HAND", 0)
    qty_s = row.get("QTY SOLD", 0)
    if qty_h > 0 and qty_h <= 5 and qty_s > 100:
        return "LOW STOCK - reorder!"
    if pd.isna(days): return "Never sold"
    d = int(days)
    if d >= 366: return "Write-off risk"
    if d >= 271: return "Critical - liquidate"
    if d >= 181: return "Near dead"
    if d >= 121: return "Deeply dormant"
    if d >= 91:  return "Dormant"
    if d >= 61:  return "At risk - act now"
    if d >= 31:  return "Slowing - monitor"
    if d >= 15:  return "Recent sale"
    if d >= 8:   return "Active"
    return "Fast mover"


def build_summary_sheet(wb, df, metrics, snap_str):
    ws = wb.active
    ws.title = "Summary Dashboard"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [3, 30, 12, 12, 14, 16, 18, 18, 24])
    write_title(ws,
        "STOCK PERFORMANCE - COMPREHENSIVE ANALYSIS REPORT",
        f"Snapshot: {snap_str}  |  Total SKUs: {metrics['total_skus']:,}  |  Stock Performance Analyzer"
    )
    kpis = [
        ("Total SKUs",       metrics["total_skus"],    "#,##0"),
        ("Active (<=30d)",   metrics["active"],        "#,##0"),
        ("Never Sold",       metrics["never_sold"],    "#,##0"),
        ("Total Revenue (N)",metrics["total_revenue"], "#,##0"),
        ("Capital at Risk",  metrics["idle_capital"],  "#,##0"),
    ]
    write_metric_cards(ws, kpis, start_row=4)

    ws.merge_cells("B7:I7")
    ws["B7"] = "STOCK HEALTH BY AGE BUCKET"
    ws["B7"].font = hf(); ws["B7"].fill = fl(MID_BLU); ws["B7"].alignment = la()
    ws.row_dimensions[7].height = 22

    hdrs = ["Age Bucket","Category","Total SKUs","With Stock","Units in Stock",
            "Capital Tied (N)","Avg Days Idle","Risk Level","Recommended Action"]
    write_table_headers(ws, hdrs, 8)

    bucket_df  = az.get_bucket_summary(df)
    risk_colors = {"Low":"375623","Medium":"7D5A00","High":"843C0C","Critical":"7F0000"}
    for ri, row in bucket_df.iterrows():
        r  = ri + 9
        bg = row["_bg"].lstrip("#")
        tc = row["_tc"].lstrip("#")
        vals = [row["Age Bucket"],row["Category"],row["Total SKUs"],row["With Stock"],
                row["Units in Stock"],row["Capital Tied (N)"],
                row["Avg Days Idle"] if row["Avg Days Idle"] else "---",
                row["Risk Level"],row["Action"]]
        fmts = [None,None,"#,##0","#,##0","#,##0","#,##0","0.0",None,None]
        for ci,(v,fm) in enumerate(zip(vals,fmts),2):
            c = ws.cell(r,ci); c.value=v; c.fill=fl(bg); c.border=bdr
            c.font = Font(name=FONT_NAME,size=10,bold=(ci==2),
                          color=tc if ci==2 else risk_colors.get(row["Risk Level"],"1F1F1F") if ci==9 else "1F1F1F")
            c.alignment = la() if ci in [2,10] else ca()
            if fm and isinstance(v,(int,float)) and not pd.isna(v): c.number_format=fm
        ws.row_dimensions[r].height = 20

    tr = len(bucket_df) + 9
    for ci in range(2,10):
        c = ws.cell(tr,ci)
        c.font=hf(col="1F1F1F"); c.fill=fl(TOT_BG); c.border=bdr; c.alignment=ca()
        if ci==2: c.value="TOTAL"
        elif ci==4: c.value=f"=SUM(E9:E{tr-1})"; c.number_format="#,##0"
        elif ci==5: c.value=f"=SUM(F9:F{tr-1})"; c.number_format="#,##0"
        elif ci==6: c.value=f"=SUM(G9:G{tr-1})"; c.number_format="#,##0"
        elif ci==7: c.value=f"=SUM(H9:H{tr-1})"; c.number_format="#,##0"
        else: c.value=""
    ws.row_dimensions[tr].height = 22


def build_detail_sheet(wb, sheet_name, title, subtitle, src_df):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [3,6,46,20,13,14,14,16,18,14,24])
    write_title(ws, title, subtitle, ncols=10)

    with_stock = src_df[src_df["QTY AT HAND"] > 0]
    avg_days   = round(src_df["DAYS_SINCE_SALE"].mean(),1) if src_df["DAYS_SINCE_SALE"].notna().any() else 0
    metrics = [
        ("Total SKUs",         len(src_df),                          "#,##0"),
        ("With Stock",         len(with_stock),                      "#,##0"),
        ("Units in Stock",     int(with_stock["QTY AT HAND"].sum()), "#,##0"),
        ("Capital Tied (N)",   int(with_stock["CAPITAL_TIED"].sum()),"#,##0"),
        ("Avg Days Since Sale",avg_days,                             "0.0"),
    ]
    write_metric_cards(ws, metrics, start_row=4, bg="F8F9FA")

    hdrs = ["#","Item Name","Last Sale Date","Days Idle","Qty Sold",
            "Qty At Hand","Unit Cost (N)","Capital Tied (N)","Barcode","Status / Flag"]
    write_table_headers(ws, hdrs, 7)

    sdf = src_df.copy()
    sdf["_s"] = sdf["DAYS_SINCE_SALE"].fillna(99999)
    sdf = sdf.sort_values(["_s","CAPITAL_TIED"],ascending=[False,False]).reset_index(drop=True)

    for idx, row in sdf.iterrows():
        r     = idx + 8
        days  = row["DAYS_SINCE_SALE"]
        qty_h = int(row["QTY AT HAND"])
        bg    = _row_bg(days, idx)
        flag  = _flag(row)
        vals  = [
            idx+1, row["ITEM NAME"], row["LAST_SALE_STR"],
            int(days) if pd.notna(days) else "---",
            int(row["QTY SOLD"]), qty_h,
            row["UNIT_COST"] if row["UNIT_COST"]>0 else 0,
            row["CAPITAL_TIED"] if row["CAPITAL_TIED"]>0 else 0,
            row.get("BARCODE_STR",""), flag,
        ]
        fmts = [None,None,None,"#,##0","#,##0","#,##0","#,##0.00","#,##0",None,None]
        for ci,(v,fm) in enumerate(zip(vals,fmts),2):
            c = ws.cell(r,ci); c.value=v
            c.font=fn(sz=9); c.fill=fl(bg); c.border=bdr
            c.alignment = la() if ci in [3,11] else ca()
            if fm and isinstance(v,(int,float)): c.number_format=fm
            if ci==7 and qty_h<0:
                c.font=Font(name=FONT_NAME,size=9,color="C00000",bold=True)
            if ci==9 and isinstance(v,(int,float)) and v>100000:
                c.font=Font(name=FONT_NAME,size=9,color="843C0C",bold=True)
        ws.row_dimensions[r].height = 15

    tr = len(sdf)+8
    write_totals_row(ws, tr, 10, sum_cols=[6,7,8,9], data_start=8)
    ws.freeze_panes = "C8"


def build_monthly_sheet(wb, df):
    ws = wb.create_sheet("Monthly Trend")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [3,6,22,16,16,18])
    write_title(ws,"MONTHLY SALES TREND","Products by last-sale month",ncols=5)
    write_table_headers(ws,["Month","Active SKUs","Units Sold","Revenue (N)","Avg Rev/SKU"],4)
    ws.row_dimensions[4].height = 20

    trend = az.get_monthly_trend(df)
    bgs   = ["EBF5E0","FFF2CC","FCE4D6","D9E8F5","E8D5F5","F2F8EC","FFEB9C","F8CBAD"]
    for ri,row in trend.iterrows():
        r   = ri+5
        bg  = bgs[ri % len(bgs)]
        rps = row["revenue"]/row["skus"] if row["skus"]>0 else 0
        for ci,(v,fm) in enumerate(zip(
            [row["MONTH_STR"],int(row["skus"]),int(row["qty_sold"]),row["revenue"],rps],
            [None,"#,##0","#,##0","#,##0","#,##0"]
        ),2):
            c=ws.cell(r,ci); c.value=v
            c.font=fn(sz=10,bold=(ci==2)); c.fill=fl(bg); c.border=bdr; c.alignment=ca()
            if fm: c.number_format=fm
        ws.row_dimensions[r].height=20

    tr=len(trend)+5
    for ci in range(2,7):
        c=ws.cell(tr,ci); c.font=hf(col="1F1F1F"); c.fill=fl(TOT_BG); c.border=bdr; c.alignment=ca()
        if ci==2: c.value="TOTAL"
        elif ci==3: c.value=f"=SUM(C5:C{tr-1})"; c.number_format="#,##0"
        elif ci==4: c.value=f"=SUM(D5:D{tr-1})"; c.number_format="#,##0"
        elif ci==5: c.value=f"=SUM(E5:E{tr-1})"; c.number_format="#,##0"
        else: c.value=""
    ws.row_dimensions[tr].height=20


def build_top_products_sheet(wb, df):
    ws = wb.create_sheet("Top Products")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws,[3,6,46,16,14,14,16,18,22])
    write_title(ws,"TOP 50 PRODUCTS - BY REVENUE","Ranked by cost-basis revenue. Red = low stock.",ncols=8)
    write_table_headers(ws,["Rank","Item Name","Qty Sold","Qty At Hand","Unit Cost (N)","Revenue (N)","Last Sale Date","Status"],4)

    top50 = az.get_top_products(df,n=50,by="REVENUE")
    for ri,row in top50.iterrows():
        r     = ri+5
        qty_h = int(row.get("QTY AT HAND",0))
        low   = qty_h<=5 and row["QTY SOLD"]>100
        bg    = "FFE0E0" if low else ("E2EFDA" if ri%2==0 else "FFFFFF")
        vals  = [ri+1,row["ITEM NAME"],int(row["QTY SOLD"]),qty_h,
                 row.get("UNIT_COST",0),row["REVENUE"],
                 row.get("LAST_SALE_STR",""),
                 "LOW STOCK" if low else row.get("STATUS","")]
        fmts  = [None,None,"#,##0","#,##0","#,##0.00","#,##0",None,None]
        for ci,(v,fm) in enumerate(zip(vals,fmts),2):
            c=ws.cell(r,ci); c.value=v
            c.font=fn(sz=9); c.fill=fl(bg); c.border=bdr
            c.alignment=la() if ci==3 else ca()
            if fm and isinstance(v,(int,float)): c.number_format=fm
        ws.row_dimensions[r].height=15
    ws.freeze_panes="C5"


def build_dead_stock_sheet(wb, df):
    ws = wb.create_sheet("Dead Stock (Never Sold)")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws,[3,6,46,14,16,18,16,24])
    write_title(ws,"DEAD STOCK - ZERO SALES, HAS PHYSICAL INVENTORY",
                "Products with QTY AT HAND > 0 that have never recorded a sale. Sorted by capital tied.",ncols=7)

    dead       = az.get_dead_stock(df)
    with_stock = dead[dead["QTY AT HAND"]>0]
    write_metric_cards(ws,[
        ("Dead Stock SKUs",   len(with_stock),                       "#,##0"),
        ("Total Units",       int(with_stock["QTY AT HAND"].sum()),  "#,##0"),
        ("Capital Tied (N)",  int(with_stock["CAPITAL_TIED"].sum()), "#,##0"),
        ("Avg Unit Cost (N)", int(with_stock["UNIT_COST"].mean()) if len(with_stock)>0 else 0,"#,##0"),
    ],start_row=4,bg="FFF2F2")
    write_table_headers(ws,["#","Item Name","Qty At Hand","Unit Cost (N)","Capital Tied (N)","Barcode","Recommendation"],7,bg="595959")

    for ri,row in dead.iterrows():
        r   = ri+8
        cap = row["CAPITAL_TIED"]
        bg  = "FFE0B2" if cap>50000 else ("FFF9E6" if cap>10000 else ("F2F2F2" if ri%2==0 else "FFFFFF"))
        vals= [ri+1,row["ITEM NAME"],int(row["QTY AT HAND"]),
               row["UNIT_COST"] if row["UNIT_COST"]>0 else 0,
               cap,row.get("BARCODE_STR",""),row.get("RECOMMENDATION","Clear or de-list")]
        fmts=[None,None,"#,##0","#,##0.00","#,##0",None,None]
        for ci,(v,fm) in enumerate(zip(vals,fmts),2):
            c=ws.cell(r,ci); c.value=v
            c.font=fn(sz=9); c.fill=fl(bg); c.border=bdr
            c.alignment=la() if ci in [3,8] else ca()
            if fm and isinstance(v,(int,float)): c.number_format=fm
        ws.row_dimensions[r].height=15

    tr=len(dead)+8
    for ci in range(2,9):
        c=ws.cell(tr,ci); c.font=hf(col="1F1F1F"); c.fill=fl(TOT_BG); c.border=bdr; c.alignment=ca()
        if ci==2: c.value="TOTAL"
        elif ci==4: c.value=f"=SUM(E8:E{tr-1})"; c.number_format="#,##0"
        elif ci==6: c.value=f"=SUM(G8:G{tr-1})"; c.number_format="#,##0"
        else: c.value=""
    ws.row_dimensions[tr].height=20
    ws.freeze_panes="C8"


def build_negative_stock_sheet(wb, df):
    neg = az.get_negative_stock(df)
    if len(neg)==0: return
    ws = wb.create_sheet("Negative Stock (Data Issue)")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws,[3,6,46,16,14,16,22])
    write_title(ws,"NEGATIVE INVENTORY - DATA INTEGRITY ALERT",
                f"{len(neg)} products show negative stock. Likely bulk-break/sachet tracking errors.",ncols=6)
    write_table_headers(ws,["#","Item Name","Qty At Hand","Qty Sold","Unit Cost (N)","Likely Cause"],4,bg="C00000")
    for ri,row in neg.iterrows():
        r   = ri+5
        bg  = "FFE0E0" if ri%2==0 else "FFF0F0"
        vals= [ri+1,row["ITEM NAME"],int(row["QTY AT HAND"]),int(row["QTY SOLD"]),
               row["UNIT_COST"] if row["UNIT_COST"]>0 else 0,"Bulk-break / sachet tracking mismatch"]
        fmts= [None,None,"#,##0","#,##0","#,##0.00",None]
        for ci,(v,fm) in enumerate(zip(vals,fmts),2):
            c=ws.cell(r,ci); c.value=v
            c.font=fn(sz=9,col="C00000" if ci==4 else "1F1F1F")
            c.fill=fl(bg); c.border=bdr
            c.alignment=la() if ci==3 else ca()
            if fm and isinstance(v,(int,float)): c.number_format=fm
        ws.row_dimensions[r].height=15
    ws.freeze_panes="C5"


def build_report(df, metrics):
    snap_str = metrics["snapshot_date"].strftime("%d/%m/%Y")
    wb = Workbook()

    build_summary_sheet(wb, df, metrics, snap_str)

    slow_df = df[df["DAYS_SINCE_SALE"].between(31,90,inclusive="both")].copy()
    build_detail_sheet(wb,"Slow Moving (31-90 Days)",
        "SLOW MOVING - Last sale 31 to 90 days ago",
        f"{len(slow_df)} products | {(slow_df['QTY AT HAND']>0).sum()} with stock", slow_df)

    dorm_df = df[df["DAYS_SINCE_SALE"].between(91,180,inclusive="both")].copy()
    build_detail_sheet(wb,"Dormant (91-180 Days)",
        "DORMANT - Last sale 91 to 180 days ago",
        f"{len(dorm_df)} products | {(dorm_df['QTY AT HAND']>0).sum()} with stock", dorm_df)

    nd_df = df[df["DAYS_SINCE_SALE"]>=181].copy()
    build_detail_sheet(wb,"Near Dead (181+ Days)",
        "NEAR DEAD - Last sale 181+ days ago",
        f"{len(nd_df)} products | {(nd_df['QTY AT HAND']>0).sum()} still have stock", nd_df)

    build_dead_stock_sheet(wb, df)

    all_stock = az.get_all_stock_by_idle(df)
    build_detail_sheet(wb,"All Stock - By Idle Days",
        "ALL IN-STOCK PRODUCTS - Ranked oldest last-sale first",
        f"{len(all_stock)} products with QTY AT HAND > 0", all_stock)

    build_top_products_sheet(wb, df)
    build_monthly_sheet(wb, df)
    build_negative_stock_sheet(wb, df)

    sold_df = df[df["LAST SALES DATE"].notna()].copy()
    build_detail_sheet(wb,"All Sold - Chronological",
        "ALL SOLD PRODUCTS - Full historical register, oldest last-sale first",
        f"{len(sold_df)} products that have ever recorded a sale", sold_df)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
